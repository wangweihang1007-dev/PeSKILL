"""Read source Excel workbooks and map rows to deal records."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TARGET_SHEETS = ("通过", "放弃", "孵化及观察跟踪类")

FIELD_DEFAULTS = {
    "project_short_name": "",
    "founded_time": "",
    "city": "",
    "expected_application": "",
    "previous_valuation": "",
    "invested_institutions": "",
    "pre_money_valuation": "",
    "current_round_amount": "",
    "financing_deadline": "",
    "main_business": "",
    "value_description": "",
    "revenue": "",
    "profit": "",
    "deal_source": "",
    "pass_status": "",
    "notes_or_rejection_reason": "",
}


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _compact(value: Any) -> str:
    return "".join(ch for ch in _clean_cell(value) if not ch.isspace())


def _normalize_sheet_name(value: str) -> str:
    return _compact(value)


def _find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        if any(_compact(cell) == "项目简称" for cell in row):
            return index
    return None


def _build_column_map(header_row: tuple[Any, ...], subheader_row: tuple[Any, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    last_main_header = ""

    for index, header in enumerate(header_row):
        main_header = _compact(header)
        sub_header = _compact(subheader_row[index] if index < len(subheader_row) else "")
        if main_header:
            last_main_header = main_header
        combined = f"{last_main_header}{sub_header}"

        if main_header == "项目简称":
            columns["project_short_name"] = index
        elif main_header == "成立时间":
            columns["founded_time"] = index
        elif main_header == "城市":
            columns["city"] = index
        elif main_header == "申报预期":
            columns["expected_application"] = index
        elif main_header == "前轮估值":
            columns["previous_valuation"] = index
        elif main_header == "已投机构":
            columns["invested_institutions"] = index
        elif "估值" in last_main_header and sub_header == "投前":
            columns["pre_money_valuation"] = index
        elif "本轮投资额" in combined:
            columns["current_round_amount"] = index
        elif main_header == "融资截止时间":
            columns["financing_deadline"] = index
        elif main_header == "主营业务":
            columns["main_business"] = index
        elif main_header.startswith("价值"):
            columns["value_description"] = index
        elif "上一年度" in last_main_header and sub_header == "收入":
            columns["revenue"] = index
        elif "上一年度" in last_main_header and sub_header == "利润":
            columns["profit"] = index
        elif main_header.startswith("项目来源"):
            columns["deal_source"] = index
        elif main_header == "是否通过":
            columns["pass_status"] = index
        elif main_header.startswith("备注") or main_header.startswith("否决原因"):
            columns["notes_or_rejection_reason"] = index

    return columns


def _looks_like_subheader(row: tuple[Any, ...]) -> bool:
    labels = {_compact(cell) for cell in row if _compact(cell)}
    subheader_markers = {"投前", "投后/或本轮投资额", "收入", "利润"}
    return bool(labels & subheader_markers)


def load_deal_records(excel_path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Load records from existing target sheets in an Excel workbook."""
    path = Path(excel_path)
    processed_sheets: list[str] = []
    records: list[dict[str, Any]] = []
    skipped_empty_name_rows = 0
    total_rows = 0

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_by_normalized_name = {_normalize_sheet_name(name): name for name in workbook.sheetnames}
        for target_sheet in TARGET_SHEETS:
            sheet_name = sheet_by_normalized_name.get(_normalize_sheet_name(target_sheet))
            if sheet_name is None:
                continue

            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            header_index = _find_header_row(rows[:20])
            if header_index is None:
                continue

            header_row = rows[header_index]
            possible_subheader_row = rows[header_index + 1] if header_index + 1 < len(rows) else ()
            subheader_row = possible_subheader_row if _looks_like_subheader(possible_subheader_row) else ()
            column_map = _build_column_map(header_row, subheader_row)
            data_start_index = header_index + 2 if subheader_row else header_index + 1

            processed_sheets.append(sheet_name)
            total_rows += max(0, len(rows) - data_start_index)

            for row_index in range(data_start_index, len(rows)):
                row = rows[row_index]
                record = dict(FIELD_DEFAULTS)
                for field, column_index in column_map.items():
                    if column_index < len(row):
                        record[field] = _clean_cell(row[column_index])
                if not record["project_short_name"]:
                    skipped_empty_name_rows += 1
                    continue

                record["source_file"] = str(path)
                record["source_sheet"] = sheet_name
                record["source_row"] = row_index + 1
                records.append(record)
    finally:
        workbook.close()

    stats = {
        "total_rows": total_rows,
        "skipped_empty_name_rows": skipped_empty_name_rows,
        "processed_sheets": processed_sheets,
    }
    return records, stats
