"""Export the current SQLite database to a reviewable Excel workbook."""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from utils import (
    DEFAULT_EXPORT_DIR,
    PROCESS_NOTES,
    connect_db,
    ensure_database,
    print_result,
    resolve_path,
    write_operation_log,
)


EXPORT_TABLES = {
    "项目主表": "project_main",
    "初筛记录表": "screening_records",
    "操作日志表": "operation_logs",
    "导入异常表": "import_errors",
}

EXPORT_COLUMN_NAMES = {
    "project_id": "项目ID",
    "normalized_project_name": "标准化项目简称",
    "project_short_name": "项目简称",
    "company_name": "公司全称",
    "founded_date": "成立时间",
    "city": "城市",
    "listing_expectation": "申报预期",
    "previous_valuation": "前轮估值",
    "invested_institutions": "已投机构",
    "pre_money_valuation": "投前估值",
    "post_money_or_investment_amount": "投后估值/本轮投资额",
    "financing_deadline": "融资截止时间",
    "main_business": "主营业务",
    "value_description": "价值说明",
    "last_year_revenue": "上一年度收入",
    "last_year_profit": "上一年度利润",
    "project_source": "项目来源",
    "intake_time": "录入时间",
    "source_raw": "项目来源和录入时间原文",
    "original_category": "原始分类",
    "current_screening_category": "当前初筛分类",
    "current_pass_status": "当前是否通过",
    "remark_or_reject_reason": "备注/否决原因",
    "special_tag": "特殊标记",
    "source_file": "来源文件",
    "source_sheet": "来源Sheet",
    "source_row": "来源行号",
    "raw_data_json": "原始行JSON",
    "created_at": "创建时间",
    "updated_at": "更新时间",
    "record_id": "初筛记录ID",
    "screening_time": "初筛时间",
    "screening_round": "初筛轮次",
    "screening_category": "初筛分类",
    "pass_status": "是否通过",
    "screening_description": "初筛说明",
    "source_type": "来源类型",
    "source_detail": "来源详情",
    "operator": "操作人",
    "is_current": "是否当前记录",
    "log_id": "日志ID",
    "operation_time": "操作时间",
    "operation_type": "操作类型",
    "user_input": "用户输入",
    "input_params": "输入参数",
    "affected_project_id": "影响项目ID",
    "affected_project_name": "影响项目简称",
    "affected_count": "影响数量",
    "modified_database": "是否修改项目数据",
    "added_screening_record": "是否新增初筛记录",
    "exported_file_path": "导出文件路径",
    "result_status": "执行状态",
    "process_note": "过程说明",
    "error_id": "异常ID",
    "import_time": "导入时间",
    "sheet_name": "Sheet名称",
    "row_number": "原始行号",
    "field_name": "字段名称",
    "raw_value": "原始值",
    "error_type": "异常类型",
    "error_message": "异常说明",
    "handling_status": "处理状态",
}


def style_export(path: Path) -> None:
    """Apply lightweight review formatting without changing database values."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.iter_cols(min_row=1, max_row=min(worksheet.max_row, 100)):
            maximum = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(maximum + 2, 10), 45)
    workbook.save(path)


def export_database(
    db_path: str | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Export all four database tables without recomputing project conclusions."""
    database_path = ensure_database(db_path)
    if output_path:
        export_path = resolve_path(output_path, DEFAULT_EXPORT_DIR)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_path = DEFAULT_EXPORT_DIR / f"project_screening_export_{timestamp}.xlsx"
    export_path.parent.mkdir(parents=True, exist_ok=True)

    with connect_db(database_path) as connection:
        counts = {
            sheet_name: connection.execute(f"SELECT COUNT(*) AS count FROM {table_name}").fetchone()["count"]
            for sheet_name, table_name in EXPORT_TABLES.items()
        }
        write_operation_log(
            connection,
            operation_type="export_database",
            input_params={"db_path": str(database_path), "output_path": str(export_path)},
            affected_count=counts["项目主表"],
            modified_database=False,
            exported_file_path=str(export_path),
            result_status="success",
            process_note=PROCESS_NOTES["export"],
        )
        frames = {
            sheet_name: pd.read_sql_query(f"SELECT * FROM {table_name}", connection).rename(
                columns=EXPORT_COLUMN_NAMES
            )
            for sheet_name, table_name in EXPORT_TABLES.items()
        }

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    style_export(export_path)

    return {
        "status": "success",
        "database_path": str(database_path),
        "export_path": str(export_path.resolve()),
        "sheet_count": len(EXPORT_TABLES),
        "record_counts": {name: len(frame) for name, frame in frames.items()},
        "export_time": datetime.now().replace(microsecond=0).isoformat(sep=" "),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="将当前项目初筛数据库导出为 Excel")
    parser.add_argument("--db-path")
    parser.add_argument("--output-path")
    args = parser.parse_args()
    result = export_database(args.db_path, args.output_path)
    print_result(result, PROCESS_NOTES["export"])


if __name__ == "__main__":
    main()
