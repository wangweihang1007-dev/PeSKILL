"""Full-refresh the screening database from the PE project Excel workbook."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from utils import (
    FIELD_MAPPING_PATH,
    PROCESS_NOTES,
    clean_header,
    connect_db,
    derive_status,
    detect_special_tag,
    ensure_database,
    json_text,
    load_yaml,
    new_id,
    normalize_project_name,
    now_iso,
    print_result,
    project_id_for,
    resolve_path,
    safe_text,
    split_source_and_time,
    write_import_error,
    write_operation_log,
    DEFAULT_DB_PATH,
)


PROJECT_COLUMNS = [
    "project_short_name",
    "company_name",
    "founded_date",
    "city",
    "listing_expectation",
    "previous_valuation",
    "invested_institutions",
    "pre_money_valuation",
    "post_money_or_investment_amount",
    "financing_deadline",
    "main_business",
    "value_description",
    "last_year_revenue",
    "last_year_profit",
    "source_raw",
    "raw_status",
    "remark_or_reject_reason",
]


def find_header_row(worksheet: Any, max_scan_rows: int = 10, max_scan_columns: int = 64) -> int | None:
    """Find the row containing the project-short-name header."""
    target = clean_header("项目简称")
    for row_number in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        for column_number in range(1, min(worksheet.max_column, max_scan_columns) + 1):
            if clean_header(worksheet.cell(row_number, column_number).value) == target:
                return row_number
    return None


def build_column_map(worksheet: Any, header_row: int, field_config: dict[str, Any]) -> tuple[dict[str, int], int]:
    """Flatten one- or two-row Excel headers and match configured aliases."""
    second_row = header_row + 1
    second_values = {
        clean_header(worksheet.cell(second_row, col).value)
        for col in range(1, min(worksheet.max_column, 64) + 1)
    }
    has_subheaders = bool(second_values & {clean_header(v) for v in ("投前", "投后/或本轮投资额", "收入", "利润")})
    data_start_row = header_row + (2 if has_subheaders else 1)

    candidates: dict[int, set[str]] = {}
    for col in range(1, min(worksheet.max_column, 64) + 1):
        parent = safe_text(worksheet.cell(header_row, col).value)
        child = safe_text(worksheet.cell(second_row, col).value) if has_subheaders else ""
        variants = {clean_header(parent), clean_header(child), clean_header(parent + child)}
        candidates[col] = {item for item in variants if item}

    mapping: dict[str, int] = {}
    for field_name, settings in field_config.get("fields", {}).items():
        aliases = {clean_header(alias) for alias in settings.get("aliases", [])}
        for col, variants in candidates.items():
            if aliases & variants:
                mapping[field_name] = col
                break
    return mapping, data_start_row


def is_year_divider(value: Any) -> bool:
    """Identify rows such as 2022/2023 that divide workbook sections."""
    text = safe_text(value)
    return bool(re.fullmatch(r"20\d{2}(?:年)?", text))


def row_has_content(worksheet: Any, row_number: int, max_columns: int = 18) -> bool:
    """Return whether a source row has any visible value."""
    return any(safe_text(worksheet.cell(row_number, col).value) for col in range(1, max_columns + 1))


def read_row(worksheet: Any, row_number: int, column_map: dict[str, int]) -> dict[str, str]:
    """读取映射字段，并将可识别的 Excel 日期序列转为日期文本。"""
    result: dict[str, str] = {}
    for field in PROJECT_COLUMNS:
        if field not in column_map:
            result[field] = ""
            continue
        value = worksheet.cell(row_number, column_map[field]).value
        if field == "founded_date" and isinstance(value, (int, float)) and 20000 <= value <= 80000:
            try:
                value = from_excel(value, worksheet.parent.epoch)
            except (TypeError, ValueError, OverflowError):
                pass
        if field == "founded_date" and hasattr(value, "date"):
            value = value.date().isoformat()
        result[field] = safe_text(value)
    return result


def insert_initial_screening(
    connection: Any,
    *,
    project_id: str,
    screening_time: str,
    source: str,
    category: str,
    pass_status: str,
    remark: str,
    sheet_name: str,
    row_number: int,
    import_time: str,
) -> str:
    """Append the source row as an immutable initial screening record."""
    connection.execute("UPDATE screening_records SET is_current = 0 WHERE project_id = ?", (project_id,))
    record_id = new_id("scr")
    connection.execute(
        """
        INSERT INTO screening_records (
            record_id, project_id, screening_time, screening_round,
            project_source, screening_category, pass_status,
            screening_description, remark_or_reject_reason,
            source_type, source_detail, operator, is_current, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        """,
        (
            record_id,
            project_id,
            screening_time,
            "Excel 初始记录",
            source,
            category,
            pass_status,
            "",
            remark,
            "项目表",
            f"{sheet_name}!{row_number}",
            "小龙虾",
            import_time,
        ),
    )
    return record_id


def full_refresh_from_excel(excel_path: str, db_path: str | None = None) -> dict[str, Any]:
    """Delete old project data and rebuild it from the latest workbook."""
    source_path = Path(excel_path).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{source_path}")
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("第一版仅支持 .xlsx 或 .xlsm 文件")

    database_path = ensure_database(db_path)
    config = load_yaml(FIELD_MAPPING_PATH)
    desired_sheets = config.get("target_sheets", [])

    # pandas validates workbook accessibility and provides the actual sheet list.
    excel_file = pd.ExcelFile(source_path, engine="openpyxl")
    actual_sheets = excel_file.sheet_names
    workbook = load_workbook(source_path, data_only=True, read_only=False)

    imported_at = now_iso()
    imported_rows = 0
    screening_records = 0
    error_count = 0
    processed_sheets: list[str] = []
    per_sheet: dict[str, int] = {}

    with connect_db(database_path) as connection:
        connection.execute("BEGIN")
        connection.execute("DELETE FROM screening_records")
        connection.execute("DELETE FROM project_main")
        connection.execute("DELETE FROM import_errors")

        for expected_sheet in desired_sheets:
            matching_sheet = next(
                (name for name in actual_sheets if re.sub(r"\s+", "", name) == re.sub(r"\s+", "", expected_sheet)),
                None,
            )
            if matching_sheet is None:
                write_import_error(
                    connection,
                    import_time=imported_at,
                    source_file=str(source_path),
                    sheet_name=expected_sheet,
                    row_number=None,
                    project_short_name="",
                    field_name="sheet",
                    raw_value="",
                    error_type="unknown_sheet",
                    error_message="目标 sheet 不存在",
                    handling_status="已跳过",
                )
                error_count += 1
                continue

            worksheet = workbook[matching_sheet]
            header_row = find_header_row(worksheet)
            if header_row is None:
                write_import_error(
                    connection,
                    import_time=imported_at,
                    source_file=str(source_path),
                    sheet_name=matching_sheet,
                    row_number=None,
                    project_short_name="",
                    field_name="header",
                    raw_value="",
                    error_type="missing_required_field",
                    error_message="未在前 10 行找到“项目简称”表头",
                    handling_status="已跳过该 sheet",
                )
                error_count += 1
                continue

            column_map, data_start_row = build_column_map(worksheet, header_row, config)
            missing_fields = [
                field_name
                for field_name, settings in config.get("fields", {}).items()
                if settings.get("required") and field_name not in column_map
            ]
            if missing_fields:
                for field_name in missing_fields:
                    write_import_error(
                        connection,
                        import_time=imported_at,
                        source_file=str(source_path),
                        sheet_name=matching_sheet,
                        row_number=header_row,
                        project_short_name="",
                        field_name=field_name,
                        raw_value="",
                        error_type="missing_required_field",
                        error_message=f"必填字段未映射：{field_name}",
                        handling_status="已跳过该 sheet",
                    )
                    error_count += 1
                continue

            processed_sheets.append(matching_sheet)
            per_sheet[matching_sheet] = 0

            for row_number in range(data_start_row, worksheet.max_row + 1):
                row = read_row(worksheet, row_number, column_map)
                project_name = row["project_short_name"]

                if not row_has_content(worksheet, row_number):
                    continue
                if is_year_divider(project_name):
                    continue
                if not project_name:
                    write_import_error(
                        connection,
                        import_time=imported_at,
                        source_file=str(source_path),
                        sheet_name=matching_sheet,
                        row_number=row_number,
                        project_short_name="",
                        field_name="project_short_name",
                        raw_value="",
                        error_type="missing_project_name",
                        error_message="项目简称为空，该行未进入项目主表",
                        handling_status="已跳过",
                    )
                    error_count += 1
                    continue

                normalized_name = normalize_project_name(project_name)
                if not normalized_name:
                    write_import_error(
                        connection,
                        import_time=imported_at,
                        source_file=str(source_path),
                        sheet_name=matching_sheet,
                        row_number=row_number,
                        project_short_name=project_name,
                        field_name="project_short_name",
                        raw_value=project_name,
                        error_type="invalid_project_name",
                        error_message="项目简称标准化后为空",
                        handling_status="已跳过",
                    )
                    error_count += 1
                    continue

                project_id = project_id_for(project_name, f"{matching_sheet}!{row_number}")
                source, intake_time = split_source_and_time(row["source_raw"])
                special_tag = detect_special_tag(worksheet.cell(row_number, column_map["project_short_name"]))
                category, pass_status = derive_status(row["raw_status"], matching_sheet, special_tag)
                screening_time = intake_time or imported_at
                connection.execute(
                    """
                    INSERT INTO project_main (
                        project_id, normalized_project_name, project_short_name,
                        company_name, founded_date, city, listing_expectation,
                        previous_valuation, invested_institutions,
                        pre_money_valuation, post_money_or_investment_amount,
                        financing_deadline, main_business, value_description,
                        last_year_revenue, last_year_profit, project_source,
                        intake_time, source_raw, original_category,
                        current_screening_category, current_pass_status,
                        remark_or_reject_reason, special_tag, source_file,
                        source_sheet, source_row, raw_data_json, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project_id,
                        normalized_name,
                        project_name,
                        row["company_name"],
                        row["founded_date"],
                        row["city"],
                        row["listing_expectation"],
                        row["previous_valuation"],
                        row["invested_institutions"],
                        row["pre_money_valuation"],
                        row["post_money_or_investment_amount"],
                        row["financing_deadline"],
                        row["main_business"],
                        row["value_description"],
                        row["last_year_revenue"],
                        row["last_year_profit"],
                        source,
                        intake_time,
                        row["source_raw"],
                        matching_sheet,
                        category,
                        pass_status,
                        row["remark_or_reject_reason"],
                        special_tag,
                        str(source_path),
                        matching_sheet,
                        row_number,
                        json_text(row),
                        imported_at,
                        imported_at,
                    ),
                )

                insert_initial_screening(
                    connection,
                    project_id=project_id,
                    screening_time=screening_time,
                    source=source or row["source_raw"],
                    category=category,
                    pass_status=pass_status,
                    remark=row["remark_or_reject_reason"],
                    sheet_name=matching_sheet,
                    row_number=row_number,
                    import_time=imported_at,
                )
                connection.execute(
                    """
                    UPDATE project_main SET
                        current_screening_category = ?,
                        current_pass_status = ?,
                        remark_or_reject_reason = CASE WHEN ? <> '' THEN ? ELSE remark_or_reject_reason END,
                        updated_at = ?
                    WHERE project_id = ?
                    """,
                    (
                        category,
                        pass_status,
                        row["remark_or_reject_reason"],
                        row["remark_or_reject_reason"],
                        imported_at,
                        project_id,
                    ),
                )
                imported_rows += 1
                screening_records += 1
                per_sheet[matching_sheet] += 1

        project_count = connection.execute("SELECT COUNT(*) AS count FROM project_main").fetchone()["count"]
        error_count = connection.execute("SELECT COUNT(*) AS count FROM import_errors").fetchone()["count"]
        write_operation_log(
            connection,
            operation_type="full_refresh",
            user_input=str(source_path),
            input_params={"excel_path": str(source_path), "processed_sheets": processed_sheets},
            affected_count=project_count,
            modified_database=True,
            added_screening_record=screening_records > 0,
            result_status="success",
            process_note=PROCESS_NOTES["import"],
        )
        connection.commit()

    return {
        "status": "success",
        "mode": "full_refresh",
        "database_path": str(database_path),
        "source_file": str(source_path),
        "available_sheets": actual_sheets,
        "processed_sheets": processed_sheets,
        "imported_source_rows": imported_rows,
        "project_count": project_count,
        "screening_record_count": screening_records,
        "per_sheet": per_sheet,
        "import_error_count": error_count,
        "import_time": imported_at,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="从 Excel 全量重建 PE 项目初筛数据库")
    parser.add_argument("excel_path", help="最新项目 Excel 文件")
    parser.add_argument("--db-path", help="SQLite 路径；默认 data/project_screening.db")
    args = parser.parse_args()
    result = full_refresh_from_excel(args.excel_path, args.db_path)
    print_result(result, PROCESS_NOTES["import"])


if __name__ == "__main__":
    main()
